import streamlit as st
import openpyxl
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
import re

# ══════════════════════════════════════════════════════════
# CONFIGURACIÓN
# ══════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Evaluación Docente - Comparativa",
    layout="wide",
)

COLOR_MAP = {
    "Sí": "#1a9641",
    "A veces": "#f39c12",
    "No": "#d7191c",
}
SCALE = ["Sí", "A veces", "No"]

DIMENSIONS = [
    "Planificacion e inicio de la sesion",
    "Desarrollo didactico de los contenidos",
    "Interaccion y participacion estudiantil",
    "Uso pedagogico de recursos tecnologicos",
    "Evaluacion formativa y retroalimentacion",
    "Gestion del tiempo y clima de aprendizaje",
]

DIMENSION_COLORS = {
    DIMENSIONS[0]: "#4e79a7",
    DIMENSIONS[1]: "#f28e2b",
    DIMENSIONS[2]: "#e15759",
    DIMENSIONS[3]: "#76b7b2",
    DIMENSIONS[4]: "#59a14f",
    DIMENSIONS[5]: "#edc948",
}

# ── Estudiante (17 preguntas tras fusionar 7+8) ────────
ESTUDIANTE_QUESTION_SHORT = {
    1: "Claridad programa y objetivos",
    2: "Organización de contenidos",
    3: "Bibliografía útil",
    4: "Puntualidad del docente",
    5: "Dominio de contenidos",
    6: "Claridad de explicaciones",
    7: "Estrategias de participación",
    8: "Relación con ejemplos prof.",
    9: "Atención de dudas",
    10: "Clima de respeto y confianza",
    11: "Orientaciones de contenido",
    12: "Coherencia de evaluaciones",
    13: "Explicación de criterios",
    14: "Retroalimentación",
    15: "Aplicación práctica",
    16: "Acceso y navegación AV",
    17: "Plataforma adecuada",
}

ESTUDIANTE_DIMENSION_MAP = {
    1: [DIMENSIONS[0]],
    2: [DIMENSIONS[0]],
    3: [DIMENSIONS[0]],
    4: [DIMENSIONS[5]],
    5: [DIMENSIONS[1]],
    6: [DIMENSIONS[1]],
    7: [DIMENSIONS[2]],
    8: [DIMENSIONS[1]],
    9: [DIMENSIONS[2]],
    10: [DIMENSIONS[5]],
    11: [DIMENSIONS[1]],
    12: [DIMENSIONS[4]],
    13: [DIMENSIONS[4]],
    14: [DIMENSIONS[4]],
    15: [DIMENSIONS[1]],
    16: [DIMENSIONS[3]],
    17: [DIMENSIONS[3]],
}

# ── Carrera (13 preguntas) ─────────────────────────────
CARRERA_QUESTION_SHORT = {
    1: "Comunicación de objetivos",
    2: "Contenidos planificados",
    3: "Orden lógico de contenidos",
    4: "Atención de inquietudes",
    5: "Consignas claras",
    6: "Recursos digitales",
    7: "Ambiente participativo",
    8: "Cumplimiento horario",
    9: "Retroalimentación",
    10: "Coordinación facilitadores",
    11: "Planificación universitaria",
    12: "Contenido en plataforma",
    13: "Seguimiento plataforma",
}

CARRERA_DIMENSION_MAP = {
    1: [DIMENSIONS[0]],
    2: [DIMENSIONS[1]],
    3: [DIMENSIONS[1]],
    4: [DIMENSIONS[2]],
    5: [DIMENSIONS[0]],
    6: [DIMENSIONS[3]],
    7: [DIMENSIONS[2]],
    8: [DIMENSIONS[5]],
    9: [DIMENSIONS[4]],
    10: [DIMENSIONS[5]],
    11: [DIMENSIONS[0]],
    12: [DIMENSIONS[3]],
    13: [DIMENSIONS[4]],
}

TIPO_ETIQUETAS = {
    "carrera": "Equipo de la Carrera",
    "estudiante": "Estudiante",
}

TIPO_COLORS = {
    "carrera": "#4e79a7",
    "estudiante": "#f28e2b",
}

# ══════════════════════════════════════════════════════════
# PARSING
# ══════════════════════════════════════════════════════════


def extract_meta_value(cell_val: str, prefix: str) -> str:
    if cell_val and cell_val.startswith(prefix):
        return cell_val[len(prefix) :].strip()
    return ""


def parse_estudiante(file_bytes: bytes, filename: str) -> dict:
    """Parsea Excel de evaluación del Estudiante al Docente.
    Escala: Excelentes→Sí, Muy Buenos→A veces, Bueno→No. Regular/Insuf. ignorados.
    Fusiona las preguntas 7 y 8 (duplicadas) promediando sus valores.
    Devuelve 17 preguntas.
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    meta = {
        "carrera": "",
        "sede": "",
        "modulo": "",
        "grupo": "",
        "docente": "",
        "fecha_inicio": "",
        "fecha_fin": "",
        "respuestas": 0,
        "archivo": filename,
        "tipo": "estudiante",
    }
    comments = []
    in_questions = False
    in_comments = False

    # Para fusionar Q7+Q8: detectar texto repetido
    seen_texts = {}

    for row in rows:
        first = str(row[0] or "").strip()

        if first.startswith("CARRERA:"):
            meta["carrera"] = extract_meta_value(first, "CARRERA:")
        elif first.startswith("SEDE:"):
            meta["sede"] = extract_meta_value(first, "SEDE:")
        elif first.startswith("MODULO:"):
            meta["modulo"] = extract_meta_value(first, "MODULO:")
        elif first.startswith("GRUPO:"):
            meta["grupo"] = extract_meta_value(first, "GRUPO:")
        elif first.startswith("DOCENTE:"):
            meta["docente"] = extract_meta_value(first, "DOCENTE:")
        elif first.startswith("FECHA INICIO:"):
            meta["fecha_inicio"] = extract_meta_value(first, "FECHA INICIO:")
        elif first.startswith("FECHA FIN:"):
            meta["fecha_fin"] = extract_meta_value(first, "FECHA FIN:")
        elif first.startswith("RESPUESTAS ENVIADAS:"):
            try:
                meta["respuestas"] = int(
                    extract_meta_value(first, "RESPUESTAS ENVIADAS:")
                )
            except:
                pass
        elif first == "Etiqueta":
            in_questions = True
            continue
        elif in_questions and not in_comments:
            pregunta = str(row[1] or "").strip()
            if "observaciones" in pregunta.lower() or "comentarios" in pregunta.lower():
                in_questions = False
                in_comments = True
                continue

            if pregunta and len(pregunta) > 5:
                num_match = re.match(r"^(\d+)[\.\-]", pregunta)
                # Clean text: remove leading number prefix for dedup
                clean_text = re.sub(r"^\d+[\.\-]\s*", "", pregunta).strip()

                def safe(val):
                    try:
                        return float(val or 0)
                    except:
                        return 0.0

                si = safe(row[2])  # Excelentes
                aveces = safe(row[4])  # Muy Buenos
                no = safe(row[6])  # Bueno

                # Check for duplicate text
                if clean_text in seen_texts:
                    idx = seen_texts[clean_text]
                    prev = st.session_state._parsing_buffer[idx]
                    prev["Sí"] = (prev["Sí"] + si) / 2
                    prev["A veces"] = (prev["A veces"] + aveces) / 2
                    prev["No"] = (prev["No"] + no) / 2
                    continue

                num = int(num_match.group(1)) if num_match else (len(seen_texts) + 1)
                q_entry = {
                    "num": num,
                    "pregunta": pregunta,
                    "short": ESTUDIANTE_QUESTION_SHORT.get(num, f"P{num}"),
                    "dimensiones": ESTUDIANTE_DIMENSION_MAP.get(num, []),
                    "Sí": si,
                    "A veces": aveces,
                    "No": no,
                }
                seen_texts[clean_text] = num
                if "_parsing_buffer" not in st.session_state:
                    st.session_state._parsing_buffer = {}
                st.session_state._parsing_buffer[num] = q_entry

        elif in_comments:
            for cell in row[2:]:
                val = str(cell or "").strip()
                if val and val.lower() not in (
                    "ninguno",
                    "ninguna",
                    "n/a",
                    "",
                    ".",
                    "-",
                    "..",
                    "ninguna ",
                    "ninguno ",
                    "ninguna observación",
                    "ninguna observacion",
                    "sin observaciones",
                    "sin comentarios",
                    "sin observacion",
                    "sin comentario",
                    "ninguna observacion ",
                    "ninguna observación ",
                    "sin observaciones.",
                    "nada",
                    "0",
                    "ningún",
                    "ningun",
                ):
                    comments.append(val)

    if not meta["docente"]:
        name_part = filename.replace(".xlsx", "").replace("_", " ")
        meta["docente"] = name_part.strip()

    # Build questions list from buffer
    questions = []
    buf = getattr(st.session_state, "_parsing_buffer", {})
    for num in sorted(buf.keys()):
        q = buf[num]
        total = q["Sí"] + q["A veces"] + q["No"]
        if total > 0:
            q["score"] = round(
                (q["Sí"] * 3 + q["A veces"] * 2 + q["No"] * 1) / total, 2
            )
        else:
            q["score"] = 0.0
        questions.append(q)

    if "_parsing_buffer" in st.session_state:
        del st.session_state._parsing_buffer

    return {"meta": meta, "questions": questions, "comments": comments}


def parse_carrera(file_bytes: bytes, filename: str) -> dict:
    """Parsea Excel de evaluación del Equipo de la Carrera al Docente.
    Escala: Excelentes→Sí, Muy Buenos→A veces, Bueno→No. Regular/Insuf. ignorados.
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    meta = {
        "carrera": "",
        "sede": "",
        "modulo": "",
        "grupo": "",
        "docente": "",
        "fecha_inicio": "",
        "fecha_fin": "",
        "respuestas": 0,
        "archivo": filename,
        "tipo": "carrera",
    }
    questions = []
    comments = []
    in_questions = False
    in_comments = False

    for row in rows:
        first = str(row[0] or "").strip()

        if first.startswith("CARRERA:"):
            meta["carrera"] = extract_meta_value(first, "CARRERA:")
        elif first.startswith("SEDE:"):
            meta["sede"] = extract_meta_value(first, "SEDE:")
        elif first.startswith("MODULO:"):
            meta["modulo"] = extract_meta_value(first, "MODULO:")
        elif first.startswith("GRUPO:"):
            meta["grupo"] = extract_meta_value(first, "GRUPO:")
        elif first.startswith("DOCENTE:"):
            meta["docente"] = extract_meta_value(first, "DOCENTE:")
        elif first.startswith("FECHA INICIO:"):
            meta["fecha_inicio"] = extract_meta_value(first, "FECHA INICIO:")
        elif first.startswith("FECHA FIN:"):
            meta["fecha_fin"] = extract_meta_value(first, "FECHA FIN:")
        elif first.startswith("RESPUESTAS ENVIADAS:"):
            try:
                meta["respuestas"] = int(
                    extract_meta_value(first, "RESPUESTAS ENVIADAS:")
                )
            except:
                pass
        elif first == "Etiqueta":
            in_questions = True
            continue
        elif in_questions and not in_comments:
            pregunta = str(row[1] or "").strip()
            if "observaciones" in pregunta.lower() or "comentarios" in pregunta.lower():
                in_questions = False
                in_comments = True
                continue

            if pregunta and len(pregunta) > 5:
                num_match = re.match(r"^(\d+)[\.\-]", pregunta)
                num = int(num_match.group(1)) if num_match else len(questions) + 1

                if num > 13:
                    continue

                def safe(val):
                    try:
                        return float(val or 0)
                    except:
                        return 0.0

                si = safe(row[2])  # Excelentes
                aveces = safe(row[4])  # Muy Buenos
                no = safe(row[6])  # Bueno

                questions.append(
                    {
                        "num": num,
                        "pregunta": pregunta,
                        "short": CARRERA_QUESTION_SHORT.get(num, f"P{num}"),
                        "dimensiones": CARRERA_DIMENSION_MAP.get(num, []),
                        "Sí": si,
                        "A veces": aveces,
                        "No": no,
                    }
                )
        elif in_comments:
            for cell in row[2:]:
                val = str(cell or "").strip()
                if val and val.lower() not in ("ninguno", "ninguna", "n/a", ""):
                    comments.append(val)

    if not meta["docente"]:
        name_part = filename.replace(".xlsx", "").replace("_", " ")
        meta["docente"] = name_part.strip()

    for q in questions:
        total = q["Sí"] + q["A veces"] + q["No"]
        if total > 0:
            q["score"] = round(
                (q["Sí"] * 3 + q["A veces"] * 2 + q["No"] * 1) / total, 2
            )
        else:
            q["score"] = 0.0

    return {"meta": meta, "questions": questions, "comments": comments}


def label_for(entry: dict) -> str:
    m = entry["meta"]
    base = m["modulo"] if m["modulo"] else m["archivo"]
    return base[:60]


# ══════════════════════════════════════════════════════════
# FUNCIONES DE VISUALIZACIÓN
# ══════════════════════════════════════════════════════════


def render_individual(entry: dict):
    """Muestra la vista de resultados para una entrada."""
    meta = entry["meta"]
    qs = entry["questions"]
    comments = entry["comments"]

    tipo_label = TIPO_ETIQUETAS.get(meta.get("tipo"), "Evaluación")

    with st.container(border=True):
        guion = "\u2014"
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.markdown("**Módulo**")
            st.markdown(
                f"<div style='word-break:break-word;white-space:normal;line-height:1.4'>{meta['modulo'] or guion}</div>",
                unsafe_allow_html=True,
            )
        with c2:
            st.markdown(f"**Tipo**")
            st.markdown(
                f"<div style='word-break:break-word;white-space:normal;line-height:1.4'>{tipo_label}</div>",
                unsafe_allow_html=True,
            )
        with c3:
            st.markdown("**Carrera**")
            st.markdown(
                f"<div style='word-break:break-word;white-space:normal;line-height:1.4'>{meta['carrera'] or guion}</div>",
                unsafe_allow_html=True,
            )
        with c4:
            st.metric("Respuestas", meta["respuestas"])
        c5, c6, c7 = st.columns(3)
        c5.caption(f"Inicio: {meta['fecha_inicio']}")
        c6.caption(f"Fin: {meta['fecha_fin']}")
        if meta.get("docente"):
            c7.markdown(f"**Docente:** {meta['docente']}")

    st.divider()

    # ── KPIs ──
    st.subheader("Indicadores Generales")
    total_si = sum(q["Sí"] for q in qs)
    total_aveces = sum(q["A veces"] for q in qs)
    total_no = sum(q["No"] for q in qs)
    grand_total = total_si + total_aveces + total_no or 1
    avg_score = sum(q["score"] for q in qs) / len(qs) if qs else 0
    pct_si = total_si / grand_total * 100
    pct_aveces = total_aveces / grand_total * 100
    pct_no = total_no / grand_total * 100

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Puntaje promedio", f"{avg_score:.2f} / 3.00")
    k2.metric("Sí (Excelentes)", f"{pct_si:.1f}%")
    k3.metric("A veces (Muy Buenos)", f"{pct_aveces:.1f}%")
    k4.metric("No (Bueno)", f"{pct_no:.1f}%")

    st.divider()

    # ── Resumen por Dimensión ──
    st.subheader("Resumen por Dimensión")
    dim_rows = []
    for dim in DIMENSIONS:
        dim_qs = [q for q in qs if dim in q.get("dimensiones", [])]
        if dim_qs:
            avg_dim = sum(q["score"] for q in dim_qs) / len(dim_qs)
            dim_rows.append({"Dimensión": dim, "Puntaje promedio": round(avg_dim, 2)})

    if dim_rows:
        df_dim = pd.DataFrame(dim_rows)
        cols_dim = st.columns(len(DIMENSIONS))
        for i, row in enumerate(dim_rows):
            cols_dim[i].metric(row["Dimensión"][:25], row["Puntaje promedio"])

        fig_dim = px.bar(
            df_dim,
            x="Dimensión",
            y="Puntaje promedio",
            color="Dimensión",
            color_discrete_map=DIMENSION_COLORS,
            text="Puntaje promedio",
            height=380,
        )
        fig_dim.update_traces(texttemplate="%{text:.2f}", textposition="outside")
        fig_dim.update_layout(
            yaxis=dict(range=[0, 3.5], title="Puntaje /3"),
            showlegend=False,
            xaxis_tickangle=-20,
            plot_bgcolor="rgba(0,0,0,0)",
            paper_bgcolor="rgba(0,0,0,0)",
        )
        st.plotly_chart(fig_dim, use_container_width=True)

    st.divider()

    # ── Tabla resumen ──
    st.subheader("Resumen por Pregunta")
    rows_sum = []
    for q in qs:
        dim_label = (
            q.get("dimensiones", ["\u2014"])[0] if q.get("dimensiones") else "\u2014"
        )
        total_q = q["Sí"] + q["A veces"] + q["No"] or 1
        rows_sum.append(
            {
                "Dimensión": dim_label,
                "Pregunta": q["short"],
                "Sí": f"{q['Sí'] / total_q * 100:.0f}%",
                "A veces": f"{q['A veces'] / total_q * 100:.0f}%",
                "No": f"{q['No'] / total_q * 100:.0f}%",
                "Puntaje (/3)": q["score"],
            }
        )
    df_sum = pd.DataFrame(rows_sum)
    st.dataframe(
        df_sum.style.background_gradient(
            subset=["Puntaje (/3)"], cmap="RdYlGn", vmin=1, vmax=3
        ),
        use_container_width=True,
        hide_index=True,
    )

    st.divider()

    # ── Barras 100% apiladas ──
    st.subheader("Distribución de Respuestas por Pregunta")
    fig_stack = go.Figure()
    for cat in SCALE:
        fig_stack.add_trace(
            go.Bar(
                name=cat,
                y=[q["short"] for q in qs],
                x=[
                    (q[cat] / (q["Sí"] + q["A veces"] + q["No"] or 1)) * 100 for q in qs
                ],
                orientation="h",
                marker_color=COLOR_MAP[cat],
                text=[
                    f"{q[cat] / (q['Sí'] + q['A veces'] + q['No'] or 1) * 100:.0f}%"
                    if q[cat] / (q["Sí"] + q["A veces"] + q["No"] or 1) * 100 > 3
                    else ""
                    for q in qs
                ],
                textposition="inside",
                insidetextanchor="middle",
            )
        )
    fig_stack.update_layout(
        barmode="stack",
        xaxis=dict(title="% de respuestas", ticksuffix="%", range=[0, 100]),
        height=400,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=180, r=20, t=60, b=40),
    )
    st.plotly_chart(fig_stack, use_container_width=True)

    # ── Barras de puntaje ──
    st.subheader("Puntaje Ponderado por Pregunta (escala 1-3)")
    df_score = pd.DataFrame(
        {
            "Pregunta": [q["short"] for q in qs],
            "Puntaje": [q["score"] for q in qs],
            "Dimensión": [
                q.get("dimensiones", ["\u2014"])[0]
                if q.get("dimensiones")
                else "\u2014"
                for q in qs
            ],
        }
    )
    fig_score = px.bar(
        df_score,
        x="Puntaje",
        y="Pregunta",
        orientation="h",
        color="Dimensión",
        color_discrete_map=DIMENSION_COLORS,
        text="Puntaje",
        height=380,
    )
    fig_score.update_traces(texttemplate="%{text:.2f}", textposition="outside")
    fig_score.update_layout(
        xaxis=dict(range=[0, 3.5], title="Puntaje promedio"),
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=180, r=80, t=20, b=40),
    )
    st.plotly_chart(fig_score, use_container_width=True)

    st.divider()

    # ── Gráficos de torta ──
    st.subheader("Detalle por Pregunta")
    if len(qs) > 0:
        for i in range(0, len(qs), 2):
            cols = st.columns(2)
            for j, q in enumerate(qs[i : i + 2]):
                labels = [s for s in SCALE if q[s] > 0]
                values = [q[s] for s in SCALE if q[s] > 0]
                fig_pie = px.pie(
                    names=labels,
                    values=values,
                    title=q["short"],
                    color=labels,
                    color_discrete_map=COLOR_MAP,
                    hole=0.42,
                    height=290,
                )
                fig_pie.update_traces(textinfo="percent+label")
                fig_pie.update_layout(
                    showlegend=False,
                    margin=dict(t=50, b=10, l=10, r=10),
                    paper_bgcolor="rgba(0,0,0,0)",
                )
                dim_label = (
                    q.get("dimensiones", ["\u2014"])[0] if q.get("dimensiones") else ""
                )
                cols[j].plotly_chart(fig_pie, use_container_width=True)
                cols[j].caption(f"{q['pregunta']}  |  _{dim_label}_")

    st.divider()

    # ── Radar ──
    st.subheader("Perfil del Docente")
    labels_r = [q["short"] for q in qs]
    scores_r = [q["score"] for q in qs]
    fig_rad = go.Figure(
        go.Scatterpolar(
            r=scores_r + [scores_r[0]],
            theta=labels_r + [labels_r[0]],
            fill="toself",
            fillcolor="rgba(26, 150, 65, 0.2)",
            line=dict(color="#1a9641", width=2),
            name="Puntaje",
        )
    )
    fig_rad.update_layout(
        polar=dict(radialaxis=dict(visible=True, range=[0, 3], tickfont=dict(size=9))),
        height=480,
        paper_bgcolor="rgba(0,0,0,0)",
    )
    st.plotly_chart(fig_rad, use_container_width=True)

    st.divider()

    # ── Comentarios ──
    st.subheader("Comentarios y Observaciones")
    if comments:
        for i, c in enumerate(comments, 1):
            st.markdown(f"**{i}.** {c}")
    else:
        st.info("No hay comentarios escritos para esta evaluación.")


def render_comparacion(entries_carrera: list, entries_estudiante: list):
    """Muestra la comparación entre evaluaciones de Carrera y Estudiante."""
    st.subheader("Comparación: Equipo de la Carrera vs Estudiante")

    if not entries_carrera and not entries_estudiante:
        st.info("Cargue datos de ambos tipos para ver la comparación.")
        return

    # Mostrar qué módulos están disponibles
    mods_carrera = {normalize_mod(e["meta"]["modulo"]): e for e in entries_carrera}
    mods_estudiante = {
        normalize_mod(e["meta"]["modulo"]): e for e in entries_estudiante
    }
    all_mods = sorted(set(list(mods_carrera.keys()) + list(mods_estudiante.keys())))

    if not all_mods:
        st.info("No hay módulos para comparar.")
        return

    sel_mod = st.selectbox("Seleccionar módulo:", all_mods)

    e_carrera = mods_carrera.get(sel_mod)
    e_estudiante = mods_estudiante.get(sel_mod)

    guion = "\u2014"
    col1, col2 = st.columns(2)
    with col1:
        if e_carrera:
            st.markdown(
                f"**Equipo de la Carrera** — {e_carrera['meta']['docente'] or guion}"
            )
            st.caption(f"{e_carrera['meta']['respuestas']} respuestas")
        else:
            st.warning("Sin evaluación del Equipo de la Carrera")
    with col2:
        if e_estudiante:
            st.markdown(f"**Estudiante** — {e_estudiante['meta']['docente'] or guion}")
            st.caption(f"{e_estudiante['meta']['respuestas']} respuestas")
        else:
            st.warning("Sin evaluación del Estudiante")

    st.divider()

    # ── Comparación por Dimensión ──
    st.markdown("#### Comparación por Dimensión")

    dim_comp = []
    for dim in DIMENSIONS:
        row = {"Dimensión": dim}
        if e_carrera:
            dim_qs = [
                q for q in e_carrera["questions"] if dim in q.get("dimensiones", [])
            ]
            row["Equipo Carrera"] = (
                round(sum(q["score"] for q in dim_qs) / len(dim_qs), 2)
                if dim_qs
                else None
            )
        if e_estudiante:
            dim_qs = [
                q for q in e_estudiante["questions"] if dim in q.get("dimensiones", [])
            ]
            row["Estudiante"] = (
                round(sum(q["score"] for q in dim_qs) / len(dim_qs), 2)
                if dim_qs
                else None
            )

        if row.get("Equipo Carrera") is not None and row.get("Estudiante") is not None:
            row["Brecha"] = round(row["Equipo Carrera"] - row["Estudiante"], 2)
        dim_comp.append(row)

    if dim_comp:
        df_dim = pd.DataFrame(dim_comp)
        mostrar_brecha = "Brecha" in df_dim.columns
        cols_show = [
            c
            for c in ["Dimensión", "Equipo Carrera", "Estudiante", "Brecha"]
            if c in df_dim.columns
        ]
        st.dataframe(
            df_dim[cols_show].style.background_gradient(
                subset=[
                    c for c in ["Equipo Carrera", "Estudiante"] if c in df_dim.columns
                ],
                cmap="RdYlGn",
                vmin=1,
                vmax=3,
            ),
            use_container_width=True,
            hide_index=True,
        )

        # Barra agrupada
        df_long = pd.DataFrame(
            [
                {
                    "Dimensión": r["Dimensión"],
                    "Evaluación": "Equipo Carrera",
                    "Puntaje": r["Equipo Carrera"],
                }
                for r in dim_comp
                if r.get("Equipo Carrera") is not None
            ]
            + [
                {
                    "Dimensión": r["Dimensión"],
                    "Evaluación": "Estudiante",
                    "Puntaje": r["Estudiante"],
                }
                for r in dim_comp
                if r.get("Estudiante") is not None
            ]
        )

        if not df_long.empty:
            fig = px.bar(
                df_long,
                x="Dimensión",
                y="Puntaje",
                color="Evaluación",
                barmode="group",
                text="Puntaje",
                height=400,
                color_discrete_map={
                    "Equipo Carrera": "#4e79a7",
                    "Estudiante": "#f28e2b",
                },
            )
            fig.update_traces(texttemplate="%{text:.2f}", textposition="outside")
            fig.update_layout(
                yaxis=dict(range=[0, 3.5], title="Puntaje /3"),
                xaxis_tickangle=-20,
                plot_bgcolor="rgba(0,0,0,0)",
                paper_bgcolor="rgba(0,0,0,0)",
            )
            st.plotly_chart(fig, use_container_width=True)

    st.divider()

    # ── Comparación por Pregunta (solo si ambos existen) ──
    if e_carrera and e_estudiante:
        st.markdown("#### Comparación por Pregunta")

        # Normalizar preguntas: agrupar por short label que exista en ambos
        q_map_carrera = {q["short"]: q["score"] for q in e_carrera["questions"]}
        q_map_estudiante = {q["short"]: q["score"] for q in e_estudiante["questions"]}
        common_qs = sorted(set(q_map_carrera.keys()) & set(q_map_estudiante.keys()))

        if common_qs:
            comp_rows = []
            for qname in common_qs:
                comp_rows.append(
                    {
                        "Pregunta": qname,
                        "Equipo Carrera": q_map_carrera[qname],
                        "Estudiante": q_map_estudiante[qname],
                        "Brecha": round(
                            q_map_carrera[qname] - q_map_estudiante[qname], 2
                        ),
                    }
                )

            df_comp = pd.DataFrame(comp_rows)
            st.dataframe(
                df_comp.style.background_gradient(
                    subset=["Equipo Carrera", "Estudiante"],
                    cmap="RdYlGn",
                    vmin=1,
                    vmax=3,
                ),
                use_container_width=True,
                hide_index=True,
            )

            # Heatmap de brechas
            fig_breach = px.bar(
                df_comp,
                x="Pregunta",
                y="Brecha",
                color="Brecha",
                color_continuous_scale=["#d7191c", "#ffffbf", "#1a9641"],
                range_color=[-1, 1],
                text="Brecha",
                height=380,
            )
            fig_breach.update_traces(texttemplate="%{text:.2f}", textposition="outside")
            fig_breach.update_layout(
                xaxis_tickangle=-25,
                plot_bgcolor="rgba(0,0,0,0)",
                paper_bgcolor="rgba(0,0,0,0)",
            )
            st.plotly_chart(fig_breach, use_container_width=True)

        # Radar comparativo
        st.markdown("#### Radar Comparativo")
        fig_rad = go.Figure()
        if e_carrera:
            scores = [q["score"] for q in e_carrera["questions"]]
            labels = [q["short"] for q in e_carrera["questions"]]
            fig_rad.add_trace(
                go.Scatterpolar(
                    r=scores + [scores[0]],
                    theta=labels + [labels[0]],
                    fill="toself",
                    name="Equipo Carrera",
                    line=dict(color="#4e79a7", width=2),
                )
            )
        if e_estudiante:
            scores = [q["score"] for q in e_estudiante["questions"]]
            labels = [q["short"] for q in e_estudiante["questions"]]
            fig_rad.add_trace(
                go.Scatterpolar(
                    r=scores + [scores[0]],
                    theta=labels + [labels[0]],
                    fill="toself",
                    name="Estudiante",
                    line=dict(color="#f28e2b", width=2),
                )
            )
        fig_rad.update_layout(
            polar=dict(radialaxis=dict(range=[0, 3])),
            height=520,
            paper_bgcolor="rgba(0,0,0,0)",
        )
        st.plotly_chart(fig_rad, use_container_width=True)

    # ── Diagnóstico rápido ──
    if e_carrera and e_estudiante:
        st.divider()
        st.markdown("#### Diagnóstico de Brechas")
        difs = []
        for dim in DIMENSIONS:
            qs_c = [
                q for q in e_carrera["questions"] if dim in q.get("dimensiones", [])
            ]
            qs_e = [
                q for q in e_estudiante["questions"] if dim in q.get("dimensiones", [])
            ]
            if qs_c and qs_e:
                avg_c = sum(q["score"] for q in qs_c) / len(qs_c)
                avg_e = sum(q["score"] for q in qs_e) / len(qs_e)
                difs.append(
                    {
                        "Dimensión": dim,
                        "Carrera": round(avg_c, 2),
                        "Estudiante": round(avg_e, 2),
                        "Diferencia": round(avg_c - avg_e, 2),
                    }
                )

        if difs:
            df_difs = pd.DataFrame(difs)
            mayor_brecha = df_difs.loc[df_difs["Diferencia"].abs().idxmax()]
            mayor_pos = df_difs.loc[df_difs["Diferencia"].idxmax()]
            mayor_neg = df_difs.loc[df_difs["Diferencia"].idxmin()]

            col_a, col_b, col_c = st.columns(3)
            col_a.metric(
                "Mayor brecha absoluta",
                f"{mayor_brecha['Dimensión'][:20]}",
                f"{mayor_brecha['Diferencia']:+.2f}",
            )
            col_b.metric(
                "A favor de Carrera",
                f"{mayor_pos['Dimensión'][:20]}",
                f"{mayor_pos['Diferencia']:+.2f}",
            )
            col_c.metric(
                "A favor de Estudiante",
                f"{mayor_neg['Dimensión'][:20]}",
                f"{mayor_neg['Diferencia']:+.2f}",
            )


def render_diagnostico(entries_carrera: list, entries_estudiante: list):
    """Dashboard consolidado con diagnóstico general."""
    st.subheader("Diagnóstico General")

    all_entries = entries_carrera + entries_estudiante
    if not all_entries:
        st.info("No hay datos cargados.")
        return

    # ── Métricas globales ──
    total_carrera = len(entries_carrera)
    total_estudiante = len(entries_estudiante)
    total_archivos = total_carrera + total_estudiante
    total_resp_carrera = sum(e["meta"]["respuestas"] for e in entries_carrera)
    total_resp_estudiante = sum(e["meta"]["respuestas"] for e in entries_estudiante)

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Archivos Carrera", total_carrera)
    col2.metric("Archivos Estudiante", total_estudiante)
    col3.metric("Resp. Carrera", total_resp_carrera)
    col4.metric("Resp. Estudiante", total_resp_estudiante)

    st.divider()

    # ── Ranking de todos los módulos ──
    st.markdown("#### Ranking de Evaluaciones")
    rank_rows = []
    for entry in all_entries:
        qs = entry["questions"]
        avg = sum(q["score"] for q in qs) / len(qs) if qs else 0
        rank_rows.append(
            {
                "Módulo": entry["meta"]["modulo"] or entry["meta"]["archivo"],
                "Tipo": TIPO_ETIQUETAS.get(entry["meta"].get("tipo", ""), "?"),
                "Docente": entry["meta"]["docente"] or "\u2014",
                "Puntaje global": round(avg, 2),
                "Respuestas": entry["meta"]["respuestas"],
            }
        )

    if rank_rows:
        df_rank = (
            pd.DataFrame(rank_rows)
            .sort_values("Puntaje global", ascending=False)
            .reset_index(drop=True)
        )
        df_rank.index += 1

        fig_rank = px.bar(
            df_rank,
            x="Módulo",
            y="Puntaje global",
            color="Tipo",
            color_discrete_map=TIPO_COLORS,
            text="Puntaje global",
            height=380,
            barmode="group",
        )
        fig_rank.update_traces(texttemplate="%{text:.2f}", textposition="outside")
        fig_rank.update_layout(
            yaxis=dict(range=[0, 3.5]),
            xaxis_tickangle=-20,
            plot_bgcolor="rgba(0,0,0,0)",
            paper_bgcolor="rgba(0,0,0,0)",
        )
        st.plotly_chart(fig_rank, use_container_width=True)
        st.dataframe(df_rank, use_container_width=True)

    st.divider()

    # ── Promedios por Dimensión (solo Carrera vs Estudiante) ──
    st.markdown("#### Promedio por Dimensión y Tipo")
    dim_data = []
    for tipo_key, entries in [
        ("carrera", entries_carrera),
        ("estudiante", entries_estudiante),
    ]:
        for dim in DIMENSIONS:
            scores = []
            for e in entries:
                dim_qs = [q for q in e["questions"] if dim in q.get("dimensiones", [])]
                if dim_qs:
                    scores.append(sum(q["score"] for q in dim_qs) / len(dim_qs))
            if scores:
                dim_data.append(
                    {
                        "Dimensión": dim,
                        "Tipo": TIPO_ETIQUETAS[tipo_key],
                        "Promedio": round(sum(scores) / len(scores), 2),
                    }
                )

    if dim_data:
        df_dim = pd.DataFrame(dim_data)
        fig_dim = px.bar(
            df_dim,
            x="Dimensión",
            y="Promedio",
            color="Tipo",
            barmode="group",
            text="Promedio",
            height=400,
            color_discrete_map=TIPO_COLORS,
        )
        fig_dim.update_traces(texttemplate="%{text:.2f}", textposition="outside")
        fig_dim.update_layout(
            yaxis=dict(range=[0, 3.5], title="Puntaje /3"),
            xaxis_tickangle=-20,
            plot_bgcolor="rgba(0,0,0,0)",
            paper_bgcolor="rgba(0,0,0,0)",
        )
        st.plotly_chart(fig_dim, use_container_width=True)

    st.divider()

    # ── Fortalezas y Áreas de Mejora ──
    st.markdown("#### Fortalezas y Áreas de Mejora")
    fortalezas = []
    debilidades = []
    for entry in all_entries:
        for q in entry["questions"]:
            if q["score"] >= 2.8:
                fortalezas.append(
                    {
                        "Módulo": entry["meta"]["modulo"] or entry["meta"]["archivo"],
                        "Tipo": TIPO_ETIQUETAS.get(entry["meta"].get("tipo", ""), "?"),
                        "Pregunta": q["short"],
                        "Puntaje": q["score"],
                    }
                )
            elif q["score"] <= 2.0:
                debilidades.append(
                    {
                        "Módulo": entry["meta"]["modulo"] or entry["meta"]["archivo"],
                        "Tipo": TIPO_ETIQUETAS.get(entry["meta"].get("tipo", ""), "?"),
                        "Pregunta": q["short"],
                        "Puntaje": q["score"],
                    }
                )

    col_a, col_b = st.columns(2)
    with col_a:
        st.markdown("**Fortalezas** (puntaje ≥ 2.8)")
        if fortalezas:
            df_f = pd.DataFrame(fortalezas).sort_values("Puntaje", ascending=False)
            st.dataframe(df_f, use_container_width=True, hide_index=True)
        else:
            st.info("Sin fortalezas destacadas.")

    with col_b:
        st.markdown("**Áreas de mejora** (puntaje ≤ 2.0)")
        if debilidades:
            df_d = pd.DataFrame(debilidades).sort_values("Puntaje")
            st.dataframe(df_d, use_container_width=True, hide_index=True)
        else:
            st.info("Sin áreas de mejora detectadas.")

    st.divider()

    # ── Estadísticas descriptivas ──
    st.markdown("#### Estadísticas Descriptivas")
    stats_rows = []
    for entry in all_entries:
        scores = [q["score"] for q in entry["questions"]]
        if scores:
            stats_rows.append(
                {
                    "Módulo": entry["meta"]["modulo"] or entry["meta"]["archivo"],
                    "Tipo": TIPO_ETIQUETAS.get(entry["meta"].get("tipo", ""), "?"),
                    "Media": round(sum(scores) / len(scores), 2),
                    "Mediana": round(pd.Series(scores).median(), 2),
                    "Mín": round(min(scores), 2),
                    "Máx": round(max(scores), 2),
                    "Std": round(pd.Series(scores).std(), 2) if len(scores) > 1 else 0,
                }
            )
    if stats_rows:
        df_stats = pd.DataFrame(stats_rows)
        st.dataframe(df_stats, use_container_width=True, hide_index=True)


def normalize_mod(name: str) -> str:
    """Normaliza nombre de módulo para matching."""
    if not name:
        return ""
    return re.sub(r"\s+", " ", name).strip().upper()


# ══════════════════════════════════════════════════════════
# SESSION STATE
# ══════════════════════════════════════════════════════════
if "data_carrera" not in st.session_state:
    st.session_state.data_carrera = {}
if "data_estudiante" not in st.session_state:
    st.session_state.data_estudiante = {}

# ══════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## Evaluación Docente")
    st.markdown("---")
    page = st.radio(
        "Navegación",
        ["Cargar Datos", "Resultados", "Comparación", "Diagnóstico General"],
        label_visibility="collapsed",
    )
    st.markdown("---")

    n_carrera = len(st.session_state.data_carrera)
    n_estudiante = len(st.session_state.data_estudiante)
    n_resp_carrera = sum(
        d["meta"]["respuestas"] for d in st.session_state.data_carrera.values()
    )
    n_resp_estudiante = sum(
        d["meta"]["respuestas"] for d in st.session_state.data_estudiante.values()
    )

    st.metric("Archivos Carrera", n_carrera)
    st.metric("Resp. Carrera", n_resp_carrera)
    st.metric("Archivos Estudiante", n_estudiante)
    st.metric("Resp. Estudiante", n_resp_estudiante)

# ══════════════════════════════════════════════════════════
# PÁGINA 1 — CARGAR DATOS
# ══════════════════════════════════════════════════════════
if page == "Cargar Datos":
    st.title("Carga de Archivos de Encuesta")
    st.markdown(
        "Sube los archivos Excel (.xlsx) exportados desde Moodle para cada tipo de evaluación."
    )

    tab1, tab2 = st.tabs(
        ["📋 Equipo de la Carrera al Docente", "🎓 Estudiante al Docente"]
    )

    with tab1:
        st.markdown("**Evaluación del Equipo de la Carrera al Docente**")
        st.caption(
            "Escala: Excelentes → Sí, Muy Buenos → A veces, Bueno → No. Regular e Insuficiente se ignoran."
        )

        uploaded_c = st.file_uploader(
            "Seleccionar archivo(s) de Carrera:",
            type=["xlsx"],
            accept_multiple_files=True,
            key="upload_carrera",
        )

        if uploaded_c:
            nuevos = 0
            for f in uploaded_c:
                if f.name not in st.session_state.data_carrera:
                    try:
                        parsed = parse_carrera(f.read(), f.name)
                        st.session_state.data_carrera[f.name] = parsed
                        nuevos += 1
                    except Exception as e:
                        st.error(f"Error en **{f.name}**: {e}")
            if nuevos:
                st.success(f"{nuevos} archivo(s) de Carrera cargado(s).")

        if st.session_state.data_carrera:
            st.markdown("### Archivos de Carrera cargados")
            for fname, entry in list(st.session_state.data_carrera.items()):
                m = entry["meta"]
                with st.container(border=True):
                    col1, col2 = st.columns([8, 1])
                    with col1:
                        st.markdown(f"**{m['modulo'] or fname}**")
                        cols_meta = st.columns(4)
                        cols_meta[0].caption(m["docente"] or "\u2014")
                        cols_meta[1].caption(m["carrera"] or "\u2014")
                        cols_meta[2].caption(m["sede"] or "\u2014")
                        cols_meta[3].caption(f"{m['respuestas']} respuestas")
                    if col2.button("Borrar", key=f"del_c_{fname}"):
                        del st.session_state.data_carrera[fname]
                        st.rerun()

    with tab2:
        st.markdown("**Evaluación del Estudiante al Docente**")
        st.caption(
            "Escala: Excelentes → Sí, Muy Buenos → A veces, Bueno → No. Preguntas duplicadas se promedian automáticamente."
        )

        uploaded_e = st.file_uploader(
            "Seleccionar archivo(s) de Estudiante:",
            type=["xlsx"],
            accept_multiple_files=True,
            key="upload_estudiante",
        )

        if uploaded_e:
            nuevos = 0
            for f in uploaded_e:
                if f.name not in st.session_state.data_estudiante:
                    try:
                        parsed = parse_estudiante(f.read(), f.name)
                        st.session_state.data_estudiante[f.name] = parsed
                        nuevos += 1
                    except Exception as e:
                        st.error(f"Error en **{f.name}**: {e}")
            if nuevos:
                st.success(f"{nuevos} archivo(s) de Estudiante cargado(s).")

        if st.session_state.data_estudiante:
            st.markdown("### Archivos de Estudiante cargados")
            for fname, entry in list(st.session_state.data_estudiante.items()):
                m = entry["meta"]
                with st.container(border=True):
                    col1, col2 = st.columns([8, 1])
                    with col1:
                        st.markdown(f"**{m['modulo'] or fname}**")
                        cols_meta = st.columns(4)
                        cols_meta[0].caption(m["docente"] or "\u2014")
                        cols_meta[1].caption(m["carrera"] or "\u2014")
                        cols_meta[2].caption(m["sede"] or "\u2014")
                        cols_meta[3].caption(f"{m['respuestas']} respuestas")
                    if col2.button("Borrar", key=f"del_e_{fname}"):
                        del st.session_state.data_estudiante[fname]
                        st.rerun()

    # Acciones globales
    st.markdown("---")
    colA, colB, colC = st.columns([3, 2, 2])
    total_files = n_carrera + n_estudiante
    if total_files > 0:
        colA.info(
            f"**{n_resp_carrera + n_resp_estudiante}** encuestados totales en **{total_files}** archivos."
        )
    if colB.button("Limpiar todo Carrera", type="secondary"):
        st.session_state.data_carrera = {}
        st.rerun()
    if colC.button("Limpiar todo Estudiante", type="secondary"):
        st.session_state.data_estudiante = {}
        st.rerun()

# ══════════════════════════════════════════════════════════
# PÁGINA 2 — RESULTADOS INDIVIDUALES
# ══════════════════════════════════════════════════════════
elif page == "Resultados":
    st.title("Resultados de Evaluación")

    # Elegir tipo
    tipo_sel = st.radio(
        "Tipo de evaluación:", ["Equipo de la Carrera", "Estudiante"], horizontal=True
    )

    if tipo_sel == "Equipo de la Carrera":
        store = st.session_state.data_carrera
        tipo_key = "carrera"
    else:
        store = st.session_state.data_estudiante
        tipo_key = "estudiante"

    if not store:
        st.warning(f"No hay datos de {tipo_sel}. Vaya a **Cargar Datos**.")
        st.stop()

    opciones = {label_for(v): k for k, v in store.items()}
    sel_label = st.selectbox(
        f"Módulo a visualizar ({tipo_sel}):", list(opciones.keys())
    )
    entry = store[opciones[sel_label]]

    render_individual(entry)

    # Exportar
    st.divider()
    st.subheader("Exportar")
    rows_exp = []
    for q in entry["questions"]:
        total_q = q["Sí"] + q["A veces"] + q["No"] or 1
        rows_exp.append(
            {
                "Pregunta": q["pregunta"],
                "Sí": q["Sí"],
                "%Sí": round(q["Sí"] / total_q * 100, 1),
                "A veces": q["A veces"],
                "%A veces": round(q["A veces"] / total_q * 100, 1),
                "No": q["No"],
                "%No": round(q["No"] / total_q * 100, 1),
                "Puntaje(/3)": q["score"],
            }
        )
    csv = pd.DataFrame(rows_exp).to_csv(index=False).encode("utf-8")
    st.download_button(
        "Descargar resumen (CSV)", csv, f"resumen_{tipo_key}.csv", "text/csv"
    )

# ══════════════════════════════════════════════════════════
# PÁGINA 3 — COMPARACIÓN
# ══════════════════════════════════════════════════════════
elif page == "Comparación":
    st.title("Comparación Cruzada")
    st.markdown(
        "Compara las evaluaciones del **Equipo de la Carrera** vs **Estudiante** para un mismo módulo."
    )

    entries_carrera = list(st.session_state.data_carrera.values())
    entries_estudiante = list(st.session_state.data_estudiante.values())

    if not entries_carrera and not entries_estudiante:
        st.warning("Cargue datos de ambos tipos para comparar.")
        st.stop()

    render_comparacion(entries_carrera, entries_estudiante)

# ══════════════════════════════════════════════════════════
# PÁGINA 4 — DIAGNÓSTICO GENERAL
# ══════════════════════════════════════════════════════════
elif page == "Diagnóstico General":
    st.title("Diagnóstico General")
    st.markdown("Dashboard consolidado de todas las evaluaciones cargadas.")

    entries_carrera = list(st.session_state.data_carrera.values())
    entries_estudiante = list(st.session_state.data_estudiante.values())

    if not entries_carrera and not entries_estudiante:
        st.warning("No hay datos cargados. Vaya primero a **Cargar Datos**.")
        st.stop()

    render_diagnostico(entries_carrera, entries_estudiante)
